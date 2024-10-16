import platform
import sys
import os
import logging
import tkinter as tk
from tkinter import filedialog, messagebox, ttk, scrolledtext
import threading
from airtest.core.api import *
import functools
import time
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill
from poco.exceptions import PocoException
from datetime import datetime
auto_setup(__file__)
from poco.drivers.android.uiautomation import AndroidUiautomationPoco
import subprocess

logging.basicConfig(level=logging.INFO)
# 全局变量
poco = None
timer_id = None
start_time = 0
elapsed_time = 0
is_processing = False

def get_adb_path():
    system = platform.system()
    if system == "Windows":
        adb_dir = os.path.join(os.path.dirname(sys.argv[0]), 'adb', 'windows')
    elif system == "Darwin":  # macOS
        adb_dir = os.path.join(os.path.dirname(sys.argv[0]), 'adb', 'mac')
    elif system == "Linux":
        adb_dir = os.path.join(os.path.dirname(sys.argv[0]), 'adb', 'linux')
    else:
        raise EnvironmentError("Unsupported operating system")

    adb_path = os.path.join(adb_dir, 'adb' + ('.exe' if system == "Windows" else ''))
    if not os.path.isfile(adb_path) or not os.access(adb_path, os.X_OK):
        raise FileNotFoundError(f"ADB binary not found or not executable at path: {adb_path}")
    return adb_path

def check_adb_device():
    adb_path = get_adb_path()
    try:
        result = subprocess.run([adb_path, "devices"], capture_output=True, text=True, check=True)
        devices = [line.split('\t')[0] for line in result.stdout.splitlines()[1:] if line.strip()]
        if devices:
            return True, devices
        else:
            return False, []
    except subprocess.CalledProcessError:
        return False, []

def initialize_poco():
    has_device, devices = check_adb_device()
    if has_device:
        adb_path = get_adb_path()
        path = os.path.dirname(sys.argv[0])
        temp_apk_path = os.path.join(path, 'pocoservice-debug.apk')
        return AndroidUiautomationPoco(use_airtest_input=True, screenshot_each_action=False, adb_path=adb_path,apk_path=temp_apk_path)
    else:
        return None

# 写入数据到 Excel 文件
def write_to_excel(input_param, success, execution_time, exception_info, time_now):
    path = os.path.dirname(sys.argv[0])
    reports_folder = os.path.join(path, 'reports')
    os.makedirs(reports_folder, exist_ok=True)
    filename = os.path.join(reports_folder, '测试结果记录.xlsx')
    wb = Workbook() if not os.path.exists(filename) else load_workbook(filename)
    ws = wb.active if wb.sheetnames else wb.create_sheet()
    if ws.max_row == 1 and ws['A1'].value is None:
        ws.cell(row=1, column=1, value="应用名称")
        ws.cell(row=1, column=2, value="成功(TRUE)/失败(FALSE)")
        ws.cell(row=1, column=3, value="耗时(s)")
        ws.cell(row=1, column=4, value="异常问题")
        ws.cell(row=1, column=5, value="时间")
        next_row = 2
    else:
        next_row = ws.max_row + 1
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    ws.cell(row=next_row, column=1, value=input_param)
    ws.cell(row=next_row, column=2, value=success)
    ws.cell(row=next_row, column=3, value=execution_time)
    ws.cell(row=next_row, column=4, value=exception_info)
    ws.cell(row=next_row, column=5, value=time_now)
    # 如果有异常信息，则给整行添加黄色背景
    if not success:
        for col in range(1, 6):
            ws.cell(row=next_row, column=col).fill = yellow_fill
    wb.save(filename)
    return f"应用名称: {input_param}, 成功: {success}, 耗时: {execution_time}s, 异常问题: {exception_info}, 时间: {time_now}"

# 异常处理装饰器
def exception_handler(func):
    @functools.wraps(func)
    def wrapper(*args, **kwargs):
        try:
            start_time = time.time()
            result = func(*args, **kwargs)
            end_time = time.time()
            execution_time = end_time - start_time
            input_param = args[0] if args else None
            time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            output = write_to_excel(input_param, True, int(execution_time), "无", time_now)
            append_output(output)
            return result
            keyevent("BACK")
        except (PocoException, Exception) as e:
            log(e, f"在执行用例 {func.__name__} 时发生异常，请查看:{e}")
            end_time = time.time()
            execution_time = end_time - start_time
            exception_info = str(e)
            input_param = args[0] if args else None
            time_now = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            error_info = "Cannot find any visible node by query UIObjectProxy of"
            if error_info in exception_info:
                output = write_to_excel(input_param, False, int(execution_time), "事项名称存疑，建议手工检查", time_now)
            else:
                output = write_to_excel(input_param, False, int(execution_time), exception_info, time_now)
            append_output(f"发生错误：{output}"+"\n")
            keyevent("BACK")

    return wrapper

# 从 Excel 文件加载数据
def load_data_xlsx(file_path, columns=None):
    wb = load_workbook(file_path)
    ws = wb.active
    if isinstance(columns, list) and all(isinstance(c, str) for c in columns):
        column_indices = {cell.value: idx for idx, cell in enumerate(ws[1], 1)}
        columns = [column_indices[c] for c in columns if c in column_indices]
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if columns is None:
            data.append([cell.strip() if isinstance(cell, str) else cell for cell in row])
        else:
            row_data = [row[idx - 1].strip() if isinstance(row[idx - 1], str) else row[idx - 1] for idx in columns]
            data.append(row_data)
    return data, len(data) + 1

# 启动应用
def startapp():
    start_app("com.demo.android.sdzwfw.activity")
    sleep(10)

# 停止应用
def stopapp():
    stop_app("com.demo.android.sdzwfw.activity")

def if_not_exist():
    if poco(name="com.demo.android.sdzwfw.activity:id/expection_tip_tv", textMatches=".*未查询到相关数据.*").exists():
        raise ValueError("未搜索到事项")

def if_repeat(item):
    """
    检查页面上是否存在完全匹配指定文本的元素，并返回结果。
    :param text: 待匹配的文本字符串
    """
    # 获取当前页面的所有可见元素
    all_elements = poco()
    elements_text = [element.get_text() for element in all_elements]
    # 检查每个元素的文本是否完全匹配指定文本
    matched_elements = [element for element in all_elements if element.get_text() == item]
    # 根据匹配的元素数量返回结果
    if len(matched_elements) > 2:
        raise ValueError("搜索出多个重复事项")
    # elif len(matched_elements) == 2:
    #     pass
    else:
        pass
        # raise ValueError("未搜索到事项")

def if_allmatch(item):
    # 只点击名称一模一样的
    poco(type="android.widget.TextView",name="com.demo.android.sdzwfw.activity:id/tv_app_name",text=item).click()

def iscompany_or_isperson():
    path = os.path.dirname(sys.argv[0])
    template_path_company = os.path.join(path, 'tpl1727667630169.png')
    template_path_person = os.path.join(path, 'tpl1727690063231.png')
    if exists(Template(template_path_company)) or exists(Template(template_path_person)):
        raise ValueError("该事项用户和已登账号类型不一致")

def if_repair():
    path = os.path.dirname(sys.argv[0])
    template_path = os.path.join(path, 'tpl1724235301671.png')
    if poco(name="com.demo.android.sdzwfw.activity:id/expection_tip_tv", textMatches=".*不便.*").exists() or exists(Template(template_path)):
        raise ValueError("事项挂维护")

def if_404():
    if poco(text="404 Not Found").exists():
        raise ValueError("404 Not Found")

def if_apitimeout():
    if poco(text="接口调用超时，请重新上传").exists():
        raise ValueError("接口调用超时")
        poco(text="确认").click()

def info_auth():
    if poco(text="信息授权",name="com.demo.android.sdzwfw.activity:id/topbar_title_tv").exists():
        poco(name="com.demo.android.sdzwfw.activity:id/iv_empower_select",type="android.widget.ImageView").click()
        poco(name="com.demo.android.sdzwfw.activity:id/tv_agree",text="同意授权").click()

def if_face():
    if poco(text="人脸识别认证",name="com.demo.android.sdzwfw.activity:id/dialog_tilte_tv").exists():
        raise ValueError("拉起人脸识别")
        poco(name="com.demo.android.sdzwfw.activity:id/dialog_negative_btn",text="取消").click()

# def if_webviewTitle(item):
#     if poco("com.demo.android.sdzwfw.activity:id/webview_title",text=item).get_text()!=item:
#         raise ValueError("巡检事项和打开的事项不一致")

# 搜索项
@exception_handler
def top_search_person(item, sbody):
    global poco
    if poco is None:
        append_output("未连接到设备，无法执行操作。")
        return
    poco(name="查询公积金", desc="查询公积金", type="android.view.View").click()
    poco(name="com.demo.android.sdzwfw.activity:id/search_et", text="查询公积金").click()
    text(item, search=True)
    if_not_exist()
    sleep(1)
    if_repeat(item)
    if_allmatch(item)
    iscompany_or_isperson()
    info_auth()
    if_face()
    if_repair()
    if_404()
    if_apitimeout()
    sleep(3)
    # if_webviewTitle(item)
    keyevent("BACK")

# 取消定时器
def cancel_timer():
    global timer_id
    if timer_id is not None:
        root.after_cancel(timer_id)
        timer_id = None
        download_button.config(state=tk.NORMAL)  # 重新启用下载按钮
        start_button.config(state=tk.NORMAL, text="开始")  # 重新启用开始按钮并更改文字
        hide_stop_button()  # 隐藏停止按钮

# 自动测试项
def autotest_items(filepath):
    global total_rows, progress_var, start_time
    stopapp()
    startapp()
    # 调用 load_data_xlsx 函数
    result = load_data_xlsx(filepath, ['站点名称', '事项名称', '服务主体'])
    # 检查返回值是否有效
    if len(result) == 2:
        dataa, total_rows = result
    else:
        raise ValueError("读取表格文件出错，请按照示例格式重新上传.")
    progress_var.set(0)
    set_progress_details(0, 0)
    for index, row in enumerate(dataa, start=1):
        if not is_processing:
            break
            # 解包数据行
        if len(row) >= 3:
            pname, item, sbody = row
        else:
            append_output("数据行不完整，请检查数据文件。")
            continue
        top_search_person(item, sbody)
        keyevent("BACK")
        update_tkinter_progress_bar(index, total_rows)
    cancel_timer()


# 追加输出到滚动文本框
def append_output(output):
    math_output.config(state=tk.NORMAL)  # 设置为可编辑
    # 如果output包含"发生错误："则标红
    if "发生错误：" in output:
        math_output.tag_config("red", foreground="red")
        math_output.insert(tk.END, output, "red")
    else:
        math_output.insert(tk.END, output + "\n")
    math_output.see(tk.END)  # 自动滚动到底部
    math_output.config(state=tk.DISABLED)  # 恢复为不可编辑


# 选择文件
def select_file():
    filepath = filedialog.askopenfilename()
    if filepath:
        file_entry.delete(0, tk.END)
        file_entry.insert(0, filepath)
        enable_start_button()
        # 重新检查设备状态
        root.after_idle(check_device_and_set_status)


# 启用开始按钮
def enable_start_button():
    global poco
    if file_entry.get():
        if poco is not None:
            start_button.config(state=tk.NORMAL, text="开始")
        else:
            start_button.config(state=tk.DISABLED, text="未连接到设备")
    else:
        start_button.config(state=tk.DISABLED, text="选择文件后开始")


# 开始处理
def start_processing():
    global is_processing
    filepath = file_entry.get()
    if filepath:
        math_output.config(state=tk.NORMAL)
        math_output.delete('1.0', tk.END)
        math_output.config(state=tk.DISABLED)

        is_processing = True
        threading.Thread(target=start_processing_and_upload, args=(filepath,)).start()
        start_button.config(state=tk.DISABLED, text="执行中，请稍后...")  # 禁用开始按钮并更改文字
        show_stop_button()  # 显示停止按钮
        stop_button.config(state=tk.NORMAL)  # 启用停止按钮
        append_output("任务开始执行...")


# 开始处理并上传
def start_processing_and_upload(filepath):
    global start_time, elapsed_time, is_processing
    start_time = time.time()
    elapsed_time = 0
    # 调用 load_data_xlsx 函数
    result = load_data_xlsx(filepath, ['站点名称', '事项名称', '服务主体'])
    # 检查返回值是否有效
    if len(result) == 2:
        dataa, total_rows = result
    else:
        raise ValueError("读取表格文件出错，请按照示例格式重新上传.")
    progress_var.set(0)
    set_progress_details(0, 0)
    update_tkinter_progress_bar(0, total_rows)
    update_elapsed_time()
    try:
        autotest_items(filepath)
    except Exception as e:
        append_output(f"发生错误：{e}")
    cancel_timer()
    is_processing = False
    hide_stop_button()
    stop_button.config(state=tk.DISABLED)
    append_output("任务执行完成。")


# 更新 Tkinter 进度条
def update_tkinter_progress_bar(current_row, total_rows):
    global elapsed_time
    if total_rows <= 1:
        progress_percent = 100
    else:
        progress_percent = (current_row / (total_rows - 1)) * 100
    progress_var.set(progress_percent)
    set_progress_details(elapsed_time, progress_percent)
    root.update_idletasks()


# 更新已用时间
def update_elapsed_time():
    global elapsed_time, timer_id
    elapsed_time = time.time() - start_time
    set_progress_details(elapsed_time, progress_var.get())
    timer_id = root.after(1000, update_elapsed_time)  # 继续定时更新


# 设置进度详情
def set_progress_details(elapsed_time, percent_complete):
    time_label.config(text=f"耗时: {elapsed_time:.2f} 秒")
    percent_label.config(text=f"{percent_complete:.2f}%")


# 下载文件
def download_file():
    path = os.path.dirname(sys.argv[0])
    reports_folder = os.path.join(path, 'reports')
    report_path = os.path.join(reports_folder, '测试结果记录.xlsx')
    messagebox.showinfo("报告下载成功", f"报告已保存到:\n{report_path}\n点击确定关闭弹窗。")


# 检查设备并在 GUI 完成初始化后设置状态
def check_device_and_set_status():
    global poco
    has_device, devices = check_adb_device()
    if has_device:
        poco = initialize_poco()
        logging.info("Device connected successfully.")
        append_output("设备连接成功，请继续操作。\n连接的设备: " + ', '.join(devices))
        enable_start_button()
    else:
        poco = None
        logging.warning("No device connected.")
        append_output("未连接到有效设备，请检查设备连接情况。")
        enable_start_button()

# 重新连接设备
def reconnect_device():
    global poco
    append_output("正在尝试连接设备...")
    poco = initialize_poco()
    if poco is not None:
        append_output("设备连接成功。")
        enable_start_button()
    else:
        append_output("设备连接失败，请检查设备连接情况。")
        enable_start_button()

# 显示停止按钮
def show_stop_button():
    stop_button.grid(row=3, column=2, sticky='ew', padx=5, pady=(5, 10))
    stop_button.config(state=tk.NORMAL)  # 启用停止按钮


# 隐藏停止按钮
def hide_stop_button():
    stop_button.grid_remove()
    stop_button.config(state=tk.DISABLED)  # 禁用停止按钮


# 停止当前任务执行
def stop_processing():
    global is_processing
    if is_processing:
        is_processing = False
        append_output("任务已停止。")
        cancel_timer()
        hide_stop_button()  # 隐藏停止按钮
        start_button.config(state=tk.NORMAL, text="开始")  # 启用开始按钮
        # math_output.config(state=tk.NORMAL)
        # math_output.delete('1.0', tk.END)
        # math_output.config(state=tk.DISABLED)


def initialize_gui():
    global root, file_entry, start_button, download_button, math_output, progress_var, time_label, percent_label, stop_button

    root = tk.Tk()
    root.title("巡检工具小助手")
    root.geometry("700x400")  # 调整窗口大小以适应新的Text控件
    root.resizable(True, True)

    style = ttk.Style()
    style.configure('TButton', justify='center')

    file_label = tk.Label(root, text="上传文件:", anchor='w', padx=10)
    file_label.grid(row=0, column=0, sticky='w')

    file_entry = tk.Entry(root, width=50)
    file_entry.grid(row=0, column=1, padx=10, pady=5, columnspan=2, sticky='ew')

    browse_button = ttk.Button(root, text="请选择", style='TButton', command=select_file)
    browse_button.grid(row=0, column=4, padx=10, columnspan=2, sticky='ew')

    refresh_button = ttk.Button(root, text="刷新连接", style='TButton', command=reconnect_device)
    refresh_button.grid(row=0, column=3, padx=10, sticky='ew')

    time_label = tk.Label(root, text="耗时: 0.00 秒", anchor='w', padx=10)
    time_label.grid(row=1, column=0, sticky='w', pady=(5, 0))

    progress_var = tk.DoubleVar(value=0)
    progress_bar = ttk.Progressbar(root, orient=tk.HORIZONTAL, length=300, mode='determinate', variable=progress_var,
                                   maximum=100)
    progress_bar.grid(row=1, column=1, columnspan=4, padx=10, pady=(0, 0), sticky='ew')

    percent_label = tk.Label(root, text="0.0%", anchor='w', padx=10)
    percent_label.grid(row=1, column=5, padx=10, pady=(0, 0), sticky='ew')

    scrollbar = ttk.Scrollbar(root)
    math_output = scrolledtext.ScrolledText(root, height=10, wrap=tk.WORD, yscrollcommand=scrollbar.set,
                                            state=tk.DISABLED)
    math_output.grid(row=2, column=0, columnspan=6, padx=10, pady=(5, 10), sticky='nsew')
    scrollbar.config(command=math_output.yview)

    start_button = ttk.Button(root, text="选择文件后开始", state=tk.DISABLED, style='TButton', command=start_processing)
    start_button.grid(row=3, column=1, sticky='ew', columnspan=1, padx=50, pady=(5, 10))

    stop_button = ttk.Button(root, text="停止", state=tk.DISABLED, style='TButton', command=stop_processing)
    stop_button.grid(row=3, column=2, sticky='ew', padx=5, pady=(5, 10))
    hide_stop_button()  # 初始状态下隐藏停止按钮

    download_button = ttk.Button(root, text="下载报告", state=tk.DISABLED, command=download_file, style='TButton')
    download_button.grid(row=3, column=3, columnspan=3, padx=10, pady=(5, 10), sticky='ew')

    for col in range(6):
        if col == 0:
            root.columnconfigure(col, weight=0, minsize=80)
        elif col == 1:
            root.columnconfigure(col, weight=1)
        else:
            root.columnconfigure(col, weight=0)

    root.rowconfigure(0, weight=0)
    root.rowconfigure(1, weight=0)
    root.rowconfigure(2, weight=1)
    root.rowconfigure(3, weight=0)

    # 在GUI完成初始化后检查设备连接状态
    root.after_idle(check_device_and_set_status)
    root.mainloop()

def main():
    """主函数，初始化GUI并在适当时候执行功能函数"""
    initialize_gui()

if __name__ == "__main__":
    main()